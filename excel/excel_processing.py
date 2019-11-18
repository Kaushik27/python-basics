import openpyxl as xl
from openpyxl.chart import Reference, BarChart


def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook["Sheet1"]
    # cell = sheet["a1"]
    # cell = sheet.cell(1, 1)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        discounted_price = cell.value * 0.9
        discounted_price_cell = sheet.cell(row, 4)
        discounted_price_cell.value = discounted_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "a10")

    workbook.save("./transactions_final.xlsx")
